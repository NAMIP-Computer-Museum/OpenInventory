import shutil
import sys
import flickrapi
import urllib.request
import openpyxl
import time
import os
from pprint import pprint
import datetime
import json

#Clé publique de l'api Flickr
FLICKR_PUBLIC = 'a3e4e529a8580f9c25c9fe07bfddb60c'

#Clé secrète de l'api Flickr
FLICKR_SECRET = '908974b10c758cde'

#ID de l'utilisateur sur lequel rechercher les photos
FLICKR_USER_ID = '197662816@N03'

#Liste d'informations supplémentaires a obtenir sur les photos
EXTRAS = 'url_o,description,tags'

#Répertoire dans lequel télécharger les photos
FOLDER = r"C:\Users\lorys\OneDrive\Documents\Informatique\2022-2023\Stage\scriptfinal\import"

def backup(flickr=""):
    #Permet de vérifier si un mock est passé pour les tests
    if(flickr==""):
        flickr = connection()
    #Récupération via l'api de toutes les photos
    userPhotos = flickr.people.getPhotos(user_id=FLICKR_USER_ID, extras = EXTRAS)
    photos = userPhotos['photos']
    savePhotos(photos)
    return photos['photo']

def backupSinceDate(date, flickr=""):
    if(flickr==""):
        flickr = connection()
    #Vérification que la date soit correcte
    try:
        datetime.datetime.strptime(date, '%d-%m-%Y')
    except ValueError:
        print("La date entrée n'est pas valide, le format doit être jj-mm-aaaa")
        time.sleep(2)
        raise
    else:
        #Récupération des photos en fonction de la date donnée
        userPhotos = flickr.people.getPhotos(user_id=FLICKR_USER_ID, extras = EXTRAS, min_upload_date = date)
        photos = userPhotos['photos']
        savePhotos(photos)
        return photos['photo']


def searchWithTags(tag="",flickr=""):
    if(flickr==""):
        flickr = connection()
    #Récupération des photos en fonction du tag entré
    userPhotos = flickr.photos.search(user_id=FLICKR_USER_ID, extras=EXTRAS, tags=tag)
    photos = userPhotos['photos']
    savePhotos(photos)
    return photos['photo']

def searchWithTagsInsideTitle(tag="",flickr=""):
    if(flickr==""):
        flickr = connection()
    #Récupération des photos en selon le tag entré. La recherche s'effectue dans tous les champs
    userPhotos = flickr.photos.search(user_id=FLICKR_USER_ID, extras=EXTRAS, text=tag)
    photos = userPhotos['photos']
    savePhotos(photos)
    return photos['photo']

def getAllPhotosByAlbum(album_name, flickr=""):
    if(flickr==""):
        flickr = connection()
    #Récupération de tous les albums de l'utilisateur
    albums = flickr.photosets.getList(user_id=FLICKR_USER_ID)
    album_id = None

    # Recherche de l'ID de l'album correspondant au nom donné
    for album in albums['photosets']['photoset']:
        if album['title']['_content'] == album_name:
            album_id = album['id']
            break
    #Si l'album n'a pas été trouvé, le script s'arrête ici
    if album_id is None:
        print("L'album n'a pas été trouvé.")
        time.sleep(2)
        return []

    # Récupération des photos de l'album
    userPhotos = flickr.photosets.getPhotos(photoset_id=album_id, extras=EXTRAS)
    photos = userPhotos['photoset']
    savePhotos(photos)
    return photos['photo']

def savePhotos(photos):
    # Création d'un fichier Excel vide
    workbook = openpyxl.Workbook()
    urls = {}
    i = 2  # Compteur pour la feuille excel
    folderExist = False

    # Sélection de la feuille active
    sheet = workbook.active

    # Ecriture des titres des en-têtes
    sheet['A1'] = "Titre"
    sheet['B1'] = "Etat"

    try:
        # Vérifie si un dossier 'photos' existe, si non il le crée
        if not FOLDER:
            if not os.path.exists('photos'):
                if __name__ == '__main__':
                    os.makedirs('photos')
                    print("Aucun répertoire indiqué, création d'un répertoire photos dans le répertoire actuel.")
                    time.sleep(1)
            else:
                if __name__ == '__main__':
                    shutil.rmtree('photos')
                    print("Un répertoire photos existe, suppression des fichiers.")
                    time.sleep(1)
        else:
            if not os.path.exists(FOLDER):
                if not os.path.exists('photos'):
                    if __name__ == '__main__':
                        os.makedirs('photos')
                        print("Le répertoire indiqué n'existe pas, création d'un répertoire photos dans le répertoire actuel.")
                        time.sleep(1)
                else:
                    if __name__ == '__main__':
                        shutil.rmtree('photos')
                        print("Un répertoire photos existe, suppression des fichiers.")
                        time.sleep(1) 
            else:
                folderExist = True
                if __name__ == '__main__':
                    shutil.rmtree(FOLDER)
                    print("Un répertoire: ",FOLDER," existe, suppression des fichiers.")
                    time.sleep(1)
                    print("Enregistrement des photos dans le répertoire: ",FOLDER)
                    time.sleep(1)

        # Parcours des photos
        for photo in photos['photo']:
            url = photo.get('url_o')
            title = photo.get('title')
            if url:
                # Vérification si le nom du fichier existe déjà dans le dictionnaire
                if title in urls:
                    # Ajout de l'url de la photo dans le dictionnaire existant
                    filepath = urls[title]
                else:
                    # Création du nouveau répertoire avec le nom de la photo
                    dir_name = None
                    tag = None
                    error = "Photo non téléchargée - Pas de tags existants"
    
                    # Récupération du tag dans le titre ou la description
                    if title:
                        if "#" in title:
                            # Divise le titre en une liste de mots
                            tags = title.split() 
                            # Compte le nombre de mots commençant par "#"
                            count = sum(1 for tag in tags if tag.startswith("#"))  
                            if count > 1 :
                                error = "Il y a trop de tags sur cette photo"
                            else:
                                tag = title.split("#")[-1].strip()
                    if not tag:
                        # Si pas de tag dans le titre récupération de la description
                        description = photo.get('description')
                        if isinstance(description, dict):
                            description = description.get('_content', '')
                        if description and "#" in description:
                            tags = description.split() 
                            count = sum(1 for tag in tags if tag.startswith("#")) 
                            if count > 1 :
                                error = "Il y a trop de tags sur cette photo"
                            else:
                                # Si il y a une description avec un tag dedans, on le récupère
                                tag = description.split("#")[-1].strip()

                    # Si un tag a été trouvé, création du répertoire correspondant
                if __name__ == '__main__':
                    if tag:
                        if folderExist == False:
                                dir_name = os.path.join('photos', tag)
                        else:
                                dir_name = os.path.join(FOLDER, tag)
                        if not os.path.exists(dir_name):
                                os.makedirs(dir_name)

                # Ajout de l'url de la photo dans le dictionnaire avec le nouveau chemin ou None si pas de tag
                urls[title] = dir_name

                # Si un répertoire a été créé, téléchargement de la photo dans le répertoire correspondant
                filepath = None
                if urls[title]:
                    # Enregistrement des photos
                    if folderExist == False:
                        filepath = os.path.join(urls[title], "photo{}.jpg".format(i))
                    else:
                        filepath = os.path.join(FOLDER, urls[title], "photo{}.jpg".format(i))

                    if __name__ == '__main__':
                        urllib.request.urlretrieve(url, filepath)

                # Ecriture des informations dans le fichier Excel avec le lien vers la photo ou une indication s'il n'y a pas de tag
                if urls[title]:
                    sheet.cell(row=i, column=1, value=title)
                    sheet.cell(row=i, column=2, value="Photo téléchargée")
                else:
                    sheet.cell(row=i, column=1, value=title)
                    sheet.cell(row=i, column=2, value=error)
                i += 1
        if __name__ == '__main__':
            print("Téléchargement des photos effectué")
            time.sleep(2)
            workbook.save("BackupResume.xlsx")
            print("Création d'un fichier de résumé terminé")
            time.sleep(2)
    except Exception as e:
        print("Une erreur est survenue pendant la sauvegarde des photos : ",e)
        time.sleep(2)


def jsonBackup(flickr=""):
    #Permet de vérifier si un mock est passé pour les tests
    if(flickr==""):
        flickr = connection()
    #Récupération de toutes les photos
    userPhotos = flickr.people.getPhotos(user_id=FLICKR_USER_ID, extras=EXTRAS)
    photos = userPhotos['photos']
    photos_data = []

    #Parcours des photos
    for photo in photos['photo']:
        photo_data = {}
        #Récupération des infos de la photo
        photo_data['title'] = photo.get('title', '')
        photo_data['url'] = photo.get('url_o', '')
        if isinstance(photo.get('description'), dict):
            photo_data['description'] = photo['description'].get('_content', '')
        else:
            photo_data['description'] = photo.get('description', '')
        #Ajout dans le json des photos les infos
        photos_data.append(photo_data)

    if __name__ == '__main__':
        #Création d'un fichier json avec les infos des photos
        with open('backup_photos.json', 'w') as f:
            json.dump(photos_data, f)
            print("Sauvegarde des photos terminée")
    return photos['photo']

def jsonBackupWithTag(tag="", flickr=""):
    if(flickr==""):
        flickr = connection()
    #Récupération de toutes les photos
    userPhotos = flickr.photos.search(user_id=FLICKR_USER_ID, extras=EXTRAS, text=tag)
    photos = userPhotos['photos']
    photos_data = []

    #Parcours des photos
    for photo in photos['photo']:
        #Récupération des infos de la photo
        photo_data = {}
        photo_data['title'] = photo.get('title', '')
        photo_data['url'] = photo.get('url_o', '')
        if isinstance(photo.get('description'), dict):
            photo_data['description'] = photo['description'].get('_content', '')
        else:
            photo_data['description'] = photo.get('description', '')
        photo_data['tags'] = photo.get('tags', '')
        #Ajout dans le json des photos les infos
        photos_data.append(photo_data)
    if __name__ == '__main__':
        with open(f'backup_photos_{tag}.json', 'w') as f:
            json.dump(photos_data, f)
            print(f"Sauvegarde des photos avec le tag '{tag}' terminée")
    return photos['photo']

def jsonBackupAlbum(album_name, flickr=""):
    #Permet de vérifier si un mock est passé pour les tests
    if(flickr==""):
        flickr = connection()
    albums = flickr.photosets.getList(user_id=FLICKR_USER_ID)
    album_id = None
    photos_data = []

    # Recherche de l'ID de l'album correspondant au nom donné
    for album in albums['photosets']['photoset']:
        if album['title']['_content'] == album_name:
            album_id = album['id']
            break

    if album_id is None:
        print("L'album n'a pas été trouvé.")
        time.sleep(2)
        return []

    # Récupération des photos de l'album
    userPhotos = flickr.photosets.getPhotos(photoset_id=album_id, extras=EXTRAS)
    photos = userPhotos['photoset']

    #Parcours des photos
    for photo in photos['photo']:
        photo_data = {}
        photo_data['title'] = photo.get('title', '')
        photo_data['url'] = photo.get('url_o', '')
        if isinstance(photo.get('description'), dict):
            photo_data['description'] = photo['description'].get('_content', '')
        else:
            photo_data['description'] = photo.get('description', '')
            
        photos_data.append(photo_data)

    if __name__ == '__main__':
        #Création du fichier json
        with open('backup_album_photos.json', 'w') as f:
            json.dump(photos_data, f)
            print("Sauvegarde des photos terminée")
    return photos['photo']

def connection():
    print("Préparation de la backup. Ne fermez pas le script une fois lancé")
    #Connection sur l'api de flickr
    flickr = flickrapi.FlickrAPI(FLICKR_PUBLIC, FLICKR_SECRET, cache=True, format='parsed-json')
    #Test pour vérifier si la connextion à pu être réalisée
    if flickr.test.login()['stat'] == 'fail':
        raise Exception("La connexion à votre profil Flickr a échoué.")
    return flickr

#Switch permettant l'affichage des options + récupération du choix de l'utilisateur
def switch(number):
    match number:
        case "1":
            backup()
        case "2":
            backupDate = input("Entrez la date souhaitée pour la backup : ")
            try:
                datetime.datetime.strptime(backupDate, '%d-%m-%Y')
            except ValueError:
                print("La date entrée n'est pas valide, le format doit être jj-mm-aaaa")
                time.sleep(2)
                print("Erreur: Date invalide. La sauvegarde depuis la date spécifiée n'a pas été effectuée.")
                return
            backupSinceDate(backupDate)
        case "3":
            tag = input("Entrez un tag pour la recherche : ")
            searchWithTags(tag)
        case "4":
            tag = input("Entrez un tag pour la recherche : ")
            searchWithTagsInsideTitle(tag)
        case "5":
            album_name = input("Entrez le nom de l'album: ")
            getAllPhotosByAlbum(album_name)
        case "6":
            jsonBackup()
        case "7":
            album_name = input("Entrez le nom de l'album: ")
            jsonBackupAlbum(album_name)
        case "8":
            tag = input("Entrez le tag pour la recherche: ")
            jsonBackupWithTag(tag)
        case "9":
            sys.exit(0)
        case _:
            print("Le chiffre entré ne correspond à aucune action !")
            time.sleep(2)

#Boucle pour relancer le script tant que l'utilisateur ne l'arrête pas
if __name__ == '__main__':
    while True:
        os.system("cls")
        print("1: Lancer la sauvegarde complète des photos depuis Flickr.")
        print("2: Lancer une sauvegarde à partir d'une date donnée.")
        print("3: Effectuer une recherche à partir d'un tag donné.")
        print("4: Effectuer une recherche à partir d'un texte donné dans n'importe quel conteneur.")
        print("5: Récupérer toutes les photos d'un album donné.")
        print("6: Sauvegarder dans un fichier JSON.")
        print("7: Sauvegarder dans un fichier JSON d'un album.")
        print("8: Sauvegarder dans un fichier JSON par tag.")
        print("9: Quitter le programme.")

        number = input("Entrez votre choix : ")
        response = switch(number)